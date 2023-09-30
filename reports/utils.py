import openpyxl
from datetime import datetime, timedelta

from robots.models import Robot
from reports.models import CreatedReports
from R4C.settings import MEDIA_ROOT, MEDIA_URL

def create_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    unique_model_values = Robot.objects.values_list('model', flat=True).distinct()

    for value in unique_model_values:
        date_week_ago = datetime.now() - timedelta(weeks=1)
        certain_model_data = Robot.objects.filter(model=value, created__gt=date_week_ago)

        if not certain_model_data:
            continue
        
        sheet = workbook.create_sheet(title=value)
    
        sheet.cell(row=1, column=1, value='Модель')
        sheet.cell(row=1, column=2, value='Версия')
        sheet.cell(row=1, column=3, value='Количество за неделю')

        serials = {}
        for obj in certain_model_data:
            serial_num = obj.serial
            if serial_num not in serials:
                serials[serial_num] = 1
            else:
                 serials[serial_num] += 1

        current_row = 2
        for serial_num, serial_count in serials.items():
            robot_version = Robot.objects.filter(serial=serial_num).first()
            sheet.cell(row=current_row, column=1, value=value)
            sheet.cell(row=current_row, column=2, value=robot_version.version)
            sheet.cell(row=current_row, column=3, value=serial_count)
            current_row += 1
        
    if "Sheet" in workbook:
            del workbook['Sheet']

    new_report = CreatedReports.objects.create()
    new_report.save()
    workbook.save(filename=f'{MEDIA_ROOT}/reports/report.xls')

    
